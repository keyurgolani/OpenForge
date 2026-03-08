"""
XLSX processor for OpenForge Knowledge System.

Processes uploaded Excel files through:
1. Text/data extraction using openpyxl
2. Metadata extraction
3. Thumbnail generation (placeholder)
4. Chunking and embedding
"""
import logging
from pathlib import Path
from uuid import UUID
from typing import Optional

from openforge.config import get_settings
from openforge.core.content_processors.base import ContentProcessor, ProcessorResult

logger = logging.getLogger("openforge.xlsx_processor")


class XlsxProcessor(ContentProcessor):
    """Process Excel spreadsheets for knowledge storage and retrieval."""

    name = "xlsx"
    supported_types = [
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]
    supported_extensions = [".xls", ".xlsx"]

    def __init__(self):
        self.settings = get_settings()

    async def process(
        self,
        file_path: str,
        workspace_id: UUID,
        knowledge_id: Optional[UUID] = None,
        **kwargs,
    ) -> ProcessorResult:
        """
        Process an Excel spreadsheet.

        Args:
            file_path: Path to the XLSX file
            workspace_id: UUID of the workspace
            knowledge_id: Optional UUID of the knowledge entry
            **kwargs: Additional options

        Returns:
            ProcessorResult with extracted text and metadata
        """
        result = ProcessorResult(success=False)

        xlsx_path = Path(file_path)
        if not xlsx_path.exists():
            result.error = f"Excel file not found: {file_path}"
            logger.error(f"Excel file not found: {file_path}")
            return result

        try:
            # Extract text using openpyxl
            text_content, metadata = await self._extract_content(xlsx_path)
            result.content = text_content
            result.extracted_text = text_content
            result.metadata = metadata

            # Generate thumbnail (placeholder for Excel files)
            result.thumbnail_path = await self._generate_thumbnail(xlsx_path, workspace_id, knowledge_id)

            result.success = True

            # Embed if knowledge_id provided
            if knowledge_id and text_content.strip():
                try:
                    from openforge.core.knowledge_processor import knowledge_processor
                    await knowledge_processor.process_knowledge(
                        knowledge_id=knowledge_id,
                        workspace_id=workspace_id,
                        content=text_content,
                        knowledge_type="xlsx",
                        title=result.ai_title,
                        tags=[],
                    )
                    result.embedded = True
                except Exception as e:
                    logger.warning(f"Failed to embed XLSX content: {e}")

        except Exception as e:
            logger.exception(f"Error processing XLSX {knowledge_id}: {e}")
            result.error = str(e)

        return result

    async def _extract_content(self, xlsx_path: Path) -> tuple[str, dict]:
        """Extract text data from Excel spreadsheet."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
            sheets_text = []
            metadata = {"sheet_count": len(wb.sheetnames), "sheets": wb.sheetnames}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text = []
                for row in ws.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    row_text = " | ".join(cells).strip().strip("|").strip()
                    if row_text:
                        rows_text.append(row_text)

                if rows_text:
                    sheets_text.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows_text))

            return "\n\n".join(sheets_text), metadata
        except ImportError:
            logger.warning("openpyxl not installed, cannot extract XLSX content")
            return "", {}
        except Exception as e:
            logger.error(f"Failed to extract XLSX content: {e}")
            return "", {}

    async def _generate_thumbnail(
        self, xlsx_path: Path, workspace_id: UUID, knowledge_id: Optional[UUID]
    ) -> Optional[str]:
        """Generate a placeholder thumbnail for Excel files."""
        if not knowledge_id:
            return None

        try:
            from PIL import Image, ImageDraw

            workspace_dir = Path(self.settings.workspace_root) / str(workspace_id)
            thumbnail_dir = workspace_dir / "thumbnails"
            thumbnail_dir.mkdir(parents=True, exist_ok=True)

            thumbnail_path = thumbnail_dir / f"{knowledge_id}.webp"

            # Create a simple placeholder image with spreadsheet-like grid
            img = Image.new("RGB", (200, 260), color=(255, 255, 255))
            draw = ImageDraw.Draw(img)

            # Draw Excel-like icon
            draw.rectangle([20, 20, 180, 240], fill=(33, 115, 70), outline=(20, 80, 50), width=2)
            draw.text((70, 110), "X", fill=(255, 255, 255))
            draw.text((40, 150), "XLSX", fill=(255, 255, 255))

            img.save(thumbnail_path, "WEBP", quality=85)
            return f"thumbnails/{knowledge_id}.webp"
        except Exception as e:
            logger.warning(f"Failed to generate XLSX thumbnail: {e}")
            return None
