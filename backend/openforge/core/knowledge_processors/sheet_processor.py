"""Sheet Processing Pipeline.

1. Read all sheets via openpyxl, convert each to markdown table
2. Extract metadata (sheet names, row/column counts per sheet)
3. Generate placeholder thumbnail
4. Chunk text → embed → store in Qdrant
"""
from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Optional
from uuid import UUID

logger = logging.getLogger("openforge.processors.sheet")


class SheetProcessor:
    """Complete sheet knowledge processing pipeline."""

    async def process(
        self,
        knowledge_id: UUID,
        file_path: str,
        workspace_id: UUID,
        db_session=None,
    ) -> dict:
        """Run the full sheet processing pipeline. Returns metadata dict."""
        result = {
            "metadata": {},
            "text": "",
            "sheet_info": [],
        }

        # ── Step 1: Extract text as markdown tables ──
        try:
            text, sheet_info = self._extract_text(file_path)
            result["text"] = text
            result["sheet_info"] = sheet_info
        except Exception as e:
            logger.warning("Sheet extraction failed for %s: %s", knowledge_id, e)

        # ── Step 2: Extract metadata ──
        try:
            result["metadata"] = self._extract_metadata(file_path, result["sheet_info"])
        except Exception as e:
            logger.warning("Sheet metadata extraction failed for %s: %s", knowledge_id, e)

        # ── Step 3: Thumbnail ──
        thumbnail_path = None
        try:
            from openforge.config import get_settings

            settings = get_settings()
            thumbnails_dir = os.path.join(settings.uploads_root, "knowledge-thumbnails")
            os.makedirs(thumbnails_dir, exist_ok=True)
            thumb_file = os.path.join(thumbnails_dir, f"{knowledge_id}.webp")
            if self._generate_sheet_placeholder(thumb_file):
                thumbnail_path = thumb_file
        except Exception as e:
            logger.warning("Sheet thumbnail generation failed for %s: %s", knowledge_id, e)

        # ── Step 4: Embed text ──
        if result["text"] and len(result["text"].strip()) >= 20:
            try:
                await self._embed_text(knowledge_id, workspace_id, result["text"])
            except Exception as e:
                logger.warning("Sheet text embedding failed for %s: %s", knowledge_id, e)

        metadata = result["metadata"]
        return {
            "thumbnail_path": thumbnail_path,
            "file_metadata": {
                "sheet_names": metadata.get("sheet_names", []),
                "sheet_details": metadata.get("sheet_details", []),
                "total_rows": metadata.get("total_rows"),
                "total_sheets": metadata.get("total_sheets"),
            },
            "content": result["text"],
            "ai_title": Path(file_path).stem.replace("_", " ").replace("-", " ").title(),
        }

    def _extract_text(self, file_path: str) -> tuple[str, list[dict]]:
        """Extract all sheets as markdown tables."""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        parts = []
        sheet_info = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                sheet_info.append({"name": sheet_name, "rows": 0, "cols": 0})
                continue

            # Filter out completely empty rows
            non_empty_rows = []
            max_cols = 0
            for row in rows:
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in cells):
                    non_empty_rows.append(cells)
                    max_cols = max(max_cols, len(cells))

            if not non_empty_rows:
                sheet_info.append({"name": sheet_name, "rows": 0, "cols": 0})
                continue

            sheet_info.append({
                "name": sheet_name,
                "rows": len(non_empty_rows),
                "cols": max_cols,
            })

            # Build markdown table
            parts.append(f"## {sheet_name}\n")

            # Limit to first 200 rows to avoid massive output
            display_rows = non_empty_rows[:200]

            # Pad cells to uniform column count
            for row in display_rows:
                while len(row) < max_cols:
                    row.append("")

            # Header row
            header = display_rows[0]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join(["---"] * max_cols) + " |")

            # Data rows
            for row in display_rows[1:]:
                parts.append("| " + " | ".join(row) + " |")

            if len(non_empty_rows) > 200:
                parts.append(f"\n*... {len(non_empty_rows) - 200} more rows truncated*\n")

            parts.append("")  # Blank line between sheets

        wb.close()

        full_text = "\n".join(parts)
        return full_text[:100000], sheet_info

    @staticmethod
    def _generate_sheet_placeholder(output_path: str) -> bool:
        """Generate a simple table-grid placeholder thumbnail for sheets."""
        try:
            from PIL import Image, ImageDraw

            width, height = 300, 200
            img = Image.new("RGB", (width, height), (30, 30, 35))
            draw = ImageDraw.Draw(img)

            rows, cols = 5, 4
            margin_x, margin_y = 40, 30
            cell_w = (width - 2 * margin_x) // cols
            cell_h = (height - 2 * margin_y) // rows
            grid_color = (52, 180, 130)
            header_color = (40, 80, 65)

            # Shade header row
            draw.rectangle(
                [(margin_x, margin_y), (width - margin_x, margin_y + cell_h)],
                fill=header_color,
            )

            # Draw grid lines
            for r in range(rows + 1):
                y = margin_y + r * cell_h
                draw.line(
                    [(margin_x, y), (width - margin_x, y)],
                    fill=grid_color,
                    width=1,
                )
            for c in range(cols + 1):
                x = margin_x + c * cell_w
                draw.line(
                    [(x, margin_y), (x, height - margin_y)],
                    fill=grid_color,
                    width=1,
                )

            img.save(output_path, "WEBP", quality=85)
            return True
        except Exception:
            return False

    def _extract_metadata(self, file_path: str, sheet_info: list[dict]) -> dict:
        """Extract sheet metadata."""
        total_rows = sum(s.get("rows", 0) for s in sheet_info)

        return {
            "sheet_names": [s["name"] for s in sheet_info],
            "sheet_details": sheet_info,
            "total_rows": total_rows,
            "total_sheets": len(sheet_info),
        }

    async def _embed_text(
        self, knowledge_id: UUID, workspace_id: UUID, text: str
    ) -> None:
        """Embed extracted text into openforge_knowledge collection."""
        from openforge.core.knowledge_processor import knowledge_processor

        await knowledge_processor.process_knowledge(
            knowledge_id=knowledge_id,
            workspace_id=workspace_id,
            content=text,
            knowledge_type="sheet",
            title=None,
            tags=[],
        )


sheet_processor = SheetProcessor()
